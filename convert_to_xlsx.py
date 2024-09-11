import openpyxl

# Carregar a planilha com o cabeçalho
wb_header = openpyxl.load_workbook('modelo.xlsx')
sheet_header = wb_header.active

# Carregar a planilha com o conteúdo
wb_content = openpyxl.load_workbook('calendario.xlsx')
sheet_content = wb_content.active

# Criar uma nova planilha para combinar cabeçalho e conteúdo
wb_combined = openpyxl.Workbook()
sheet_combined = wb_combined.active

# Copiar o cabeçalho para a nova planilha
for row in sheet_header.iter_rows(values_only=True):
    sheet_combined.append(row)

# Copiar o conteúdo para a nova planilha, começando após o cabeçalho
for row in sheet_content.iter_rows(values_only=True):
    sheet_combined.append(row)

# Salvar a nova planilha com o cabeçalho e o conteúdo combinados
wb_combined.save('combinada.xlsx')